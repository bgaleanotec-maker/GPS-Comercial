# Ruta: SST/app/analytics/export_utils.py
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import pytz
from io import BytesIO

def generate_dashboard_excel(data):
    """
    Genera un archivo Excel completo con todos los datos del dashboard de analítica.
    
    Args:
        data: Diccionario con todos los datos del dashboard
    
    Returns:
        BytesIO: Archivo Excel en memoria
    """
    wb = Workbook()
    colombia_tz = pytz.timezone('America/Bogota')
    
    # === HOJA 1: RESUMEN GENERAL ===
    ws_summary = wb.active
    ws_summary.title = "Resumen General"
    
    # Estilos generales
    title_font = Font(bold=True, size=14, color="1F2937")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título del reporte
    ws_summary['A1'] = f"Reporte de Analítica Comercial - {datetime.now(colombia_tz).strftime('%d/%m/%Y %I:%M %p')}"
    ws_summary['A1'].font = Font(bold=True, size=16, color="4F46E5")
    ws_summary.merge_cells('A1:D1')
    
    # Periodo del reporte
    ws_summary['A3'] = "Periodo:"
    ws_summary['B3'] = f"{data.get('date_from', 'N/A')} - {data.get('date_to', 'N/A')}"
    ws_summary['A3'].font = Font(bold=True)
    
    # Métricas principales
    ws_summary['A5'] = "Métricas Generales"
    ws_summary['A5'].font = title_font
    
    metrics = [
        ["Total Empleados", data.get('total_employees', 0)],
        ["Total Aliados", data.get('total_allies', 0)],
        ["Visitas en Periodo", data.get('total_visits', 0)],
        ["Aliados Activos", data.get('active_allies', 0)],
        ["Promedio Visitas por Aliado", f"{data.get('avg_visits', 0):.1f}"],
    ]
    
    row = 6
    for metric, value in metrics:
        ws_summary[f'A{row}'] = metric
        ws_summary[f'B{row}'] = value
        ws_summary[f'A{row}'].font = Font(bold=True)
        ws_summary[f'A{row}'].border = border
        ws_summary[f'B{row}'].border = border
        row += 1
    
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 20
    
    # === HOJA 2: ANÁLISIS POR EMPLEADO ===
    ws_employees = wb.create_sheet("Por Empleado")
    
    # Encabezados
    headers_emp = ["Empleado", "Vehículo", "Km Hoy", "Km Mes", "Km Total", "Visitas", "Infracciones"]
    for col, header in enumerate(headers_emp, 1):
        cell = ws_employees.cell(1, col, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
    
    # Datos de empleados
    for idx, emp_data in enumerate(data.get('employees_data', []), 2):
        ws_employees.cell(idx, 1, emp_data.get('name', 'N/A')).border = border
        ws_employees.cell(idx, 2, emp_data.get('vehicle', 'N/A')).border = border
        ws_employees.cell(idx, 3, f"{emp_data.get('km_today', 0):.2f}").border = border
        ws_employees.cell(idx, 4, f"{emp_data.get('km_month', 0):.2f}").border = border
        ws_employees.cell(idx, 5, f"{emp_data.get('km_total', 0):.2f}").border = border
        ws_employees.cell(idx, 6, emp_data.get('visits', 0)).border = border
        ws_employees.cell(idx, 7, emp_data.get('infractions', 0)).border = border
    
    # Ajustar anchos
    for col in range(1, 8):
        ws_employees.column_dimensions[get_column_letter(col)].width = 18
    
    # === HOJA 3: ANÁLISIS POR ALIADO ===
    ws_allies = wb.create_sheet("Por Aliado")
    
    # Encabezados
    headers_ally = ["Aliado", "Categoría", "Visitas (Periodo)", "Empleados Únicos", "Última Visita", "Estado"]
    for col, header in enumerate(headers_ally, 1):
        cell = ws_allies.cell(1, col, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
    
    # Datos de aliados
    for idx, ally_data in enumerate(data.get('allies_data', []), 2):
        ws_allies.cell(idx, 1, ally_data.get('name', 'N/A')).border = border
        ws_allies.cell(idx, 2, ally_data.get('category', '-')).border = border
        ws_allies.cell(idx, 3, ally_data.get('visits_period', 0)).border = border
        ws_allies.cell(idx, 4, ally_data.get('unique_employees', 0)).border = border
        
        last_visit = ally_data.get('last_visit', '-')
        ws_allies.cell(idx, 5, last_visit).border = border
        
        estado = ally_data.get('status', 'Inactivo')
        cell_status = ws_allies.cell(idx, 6, estado)
        cell_status.border = border
        if estado == 'Activo':
            cell_status.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
        else:
            cell_status.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    
    # Ajustar anchos
    ws_allies.column_dimensions['A'].width = 30
    ws_allies.column_dimensions['B'].width = 20
    ws_allies.column_dimensions['C'].width = 18
    ws_allies.column_dimensions['D'].width = 18
    ws_allies.column_dimensions['E'].width = 20
    ws_allies.column_dimensions['F'].width = 15
    
    # === HOJA 4: DETALLE DE VISITAS ===
    ws_visits = wb.create_sheet("Detalle Visitas")
    
    # Encabezados
    headers_visits = ["Fecha/Hora", "Empleado", "Vehículo", "Aliado", "Tipo", "Categoría", "Observaciones"]
    for col, header in enumerate(headers_visits, 1):
        cell = ws_visits.cell(1, col, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
    
    # Datos de visitas
    for idx, visit in enumerate(data.get('visits_detail', []), 2):
        timestamp = visit.timestamp.astimezone(colombia_tz).strftime('%d/%m/%Y %I:%M %p')
        ws_visits.cell(idx, 1, timestamp).border = border
        ws_visits.cell(idx, 2, visit.employee_name or 'N/A').border = border
        ws_visits.cell(idx, 3, visit.device_name or 'N/A').border = border
        ws_visits.cell(idx, 4, visit.ally.name).border = border
        
        tipo = "Manual" if visit.is_manual else "Automática"
        cell_tipo = ws_visits.cell(idx, 5, tipo)
        cell_tipo.border = border
        if visit.is_manual:
            cell_tipo.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
        else:
            cell_tipo.fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
        
        ws_visits.cell(idx, 6, visit.category or '-').border = border
        ws_visits.cell(idx, 7, visit.observations or '-').border = border
    
    # Ajustar anchos
    ws_visits.column_dimensions['A'].width = 20
    ws_visits.column_dimensions['B'].width = 25
    ws_visits.column_dimensions['C'].width = 20
    ws_visits.column_dimensions['D'].width = 30
    ws_visits.column_dimensions['E'].width = 15
    ws_visits.column_dimensions['F'].width = 20
    ws_visits.column_dimensions['G'].width = 40
    
    # Guardar en memoria
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file